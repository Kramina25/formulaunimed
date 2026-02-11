import os
import sys
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

# =========================
# CONFIG
# =========================
def get_base_dir() -> str:
    # Quando vira .exe (PyInstaller), sys.executable é o caminho do executável
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    # Quando roda como .py normal
    return os.path.dirname(os.path.abspath(__file__))

PASTA = get_base_dir()
SAIDA = os.path.join(PASTA, "Resumo_Cirurgias.xlsx")


PREFIXOS = {
    "65641": "MCN",
    "59153": "DBA",
    "102410": "JBM",
    "109630": "GRM",
    "105004": "AGL",
}

# Dropdown (11 cirurgiões)
CIRURGIOES = ["RR", "MSM", "MCN", "DBA", "EBS", "JBM", "KR", "GRM", "GSJ", "AGL", "JFC"]

# Colunas que existem no Resumo (pivot) e têm "Total Unimed"
COLUNAS_RESUMO = ["MCN", "DBA", "JBM", "GRM", "AGL"]


def medico_por_arquivo(nome_arquivo: str) -> str:
    for prefixo, sigla in PREFIXOS.items():
        if nome_arquivo.startswith(prefixo):
            return sigla
    return ""


def autofit_worksheet(ws, min_width_numeric=18, max_width=60, currency_cols_from=2):
    """Ajusta largura de todas as colunas baseado no conteúdo."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        width = max_len + 3
        if col_idx >= currency_cols_from:
            width = max(width, min_width_numeric)
        ws.column_dimensions[col_letter].width = min(width, max_width)


# =========================
# LEITURA DOS ARQUIVOS
# =========================
arquivos = [
    f for f in os.listdir(PASTA)
    if f.lower().endswith(".xlsx")
    and not f.startswith("~$")
    and f != os.path.basename(SAIDA)
]

# DEBUG (para ver onde o exe está procurando)
print(f"📁 Pasta usada: {PASTA}")
print(f"📄 Arquivos .xlsx encontrados: {len(arquivos)}")
for f in arquivos:
    print(" -", f)


# DEBUG (para ver onde o exe está procurando)
print(f"📁 Pasta usada: {PASTA}")
print(f"📄 Arquivos .xlsx encontrados: {len(arquivos)}")
for f in arquivos:
    print(" -", f)

resumo_total = []

for arquivo in arquivos:
    caminho = os.path.join(PASTA, arquivo)
    try:
        xls = pd.ExcelFile(caminho)
        if "Guia de Serviços" not in xls.sheet_names or "Extrato" not in xls.sheet_names:
            continue

        servicos = pd.read_excel(xls, sheet_name="Guia de Serviços", header=None)
        extrato = pd.read_excel(xls, sheet_name="Extrato", header=None)

        # Filtro: coluna J (índice 9) != "CON"
        if servicos.shape[1] <= 9:
            continue

        filtro = servicos[servicos[9].astype(str) != "CON"]

        # Coluna B (índice 1) = códigos/itens
        if servicos.shape[1] <= 6 or servicos.shape[1] <= 1:
            continue
        if extrato.shape[1] <= 19 or extrato.shape[1] <= 5:
            continue

        b_unicos = filtro[1].dropna().unique()

        cirurgias = pd.DataFrame({"B": b_unicos})
        cirurgias["A"] = cirurgias["B"].map(dict(zip(servicos[1], servicos[6])))
        cirurgias["C"] = cirurgias["B"].map(dict(zip(extrato[5], extrato[19])))
        cirurgias["C"] = pd.to_numeric(cirurgias["C"], errors="coerce").fillna(0.0)
        cirurgias["Medico"] = medico_por_arquivo(arquivo)

        cirurgias = cirurgias[["A", "C", "Medico"]]
        resumo_total.append(cirurgias)

    except Exception as e:
        print(f"Erro em {arquivo}: {e}")

if not resumo_total:
    raise SystemExit("⚠️ Nenhum dado processado. Verifique se os arquivos têm as abas/colunas esperadas.")

base_df = pd.concat(resumo_total, ignore_index=True)

# =========================
# RESUMO (PIVOT)
# =========================
pivot = pd.pivot_table(
    base_df,
    index="A",
    columns="Medico",
    values="C",
    aggfunc="sum",
    fill_value=0.0
)

ordem = ["MCN", "DBA", "JBM", "GRM", "AGL"]
pivot = pivot[[c for c in ordem if c in pivot.columns]]

pivot["Total"] = pivot.sum(axis=1)

total_row = pivot.sum(axis=0)
total_row.name = "Total"

pivot_final = pd.concat([pivot, total_row.to_frame().T])
pivot_final.index.name = "Paciente"
pivot_final["Cirurgião"] = ""

# ✅ Reordenar colunas no Resumo: Paciente | Cirurgião | MCN.. | Total
colunas_novas = (
    ["Cirurgião"] +
    [c for c in ordem if c in pivot_final.columns] +
    ["Total"]
)
pivot_final = pivot_final[colunas_novas]


# =========================
# EXPORTAÇÃO + DROPDOWN + TOTAIS
# =========================
with pd.ExcelWriter(SAIDA, engine="openpyxl") as writer:
    # Abas principais
    base_df.to_excel(writer, sheet_name="Base", index=False)
    pivot_final.to_excel(writer, sheet_name="Resumo")

    wb = writer.book
    ws_base = writer.sheets["Base"]
    ws_resumo = writer.sheets["Resumo"]

    # Fixar cabeçalho no Resumo
    ws_resumo.freeze_panes = "A2"

    # -------------------------
    # Aba Listas (dropdown)
    # -------------------------
    if "Listas" in wb.sheetnames:
        ws_listas = wb["Listas"]
        ws_listas.delete_rows(1, ws_listas.max_row)
    else:
        ws_listas = wb.create_sheet("Listas")

    ws_listas["A1"] = "Cirurgião"
    ws_listas["A1"].font = Font(bold=True)
    for i, nome in enumerate(CIRURGIOES, start=2):
        ws_listas[f"A{i}"] = nome

    # -------------------------
    # Descobrir colunas no Resumo
    # -------------------------
    headers = [cell.value for cell in ws_resumo[1]]

    col_total_idx = headers.index("Total") + 1
    col_cir_idx = headers.index("Cirurgião") + 1

    total_col_letter = get_column_letter(col_total_idx)
    cir_col_letter = get_column_letter(col_cir_idx)

    last_row_resumo = ws_resumo.max_row          # linha final = "Total"
    data_last_row = max(2, last_row_resumo - 1)  # até antes do "Total"

    # -------------------------
    # Dropdown na coluna Cirurgião (aba Resumo)
    # -------------------------
    dv = DataValidation(
        type="list",
        formula1=f"=Listas!$A$2:$A${len(CIRURGIOES)+1}",
        allow_blank=True
    )
    ws_resumo.add_data_validation(dv)
    dv.add(f"{cir_col_letter}2:{cir_col_letter}{data_last_row}")

    # -------------------------
    # Aba Totais Cirurgião
    # -------------------------
    if "Totais Cirurgião" in wb.sheetnames:
        ws_tot = wb["Totais Cirurgião"]
        ws_tot.delete_rows(1, ws_tot.max_row)
    else:
        ws_tot = wb.create_sheet("Totais Cirurgião")

    ws_tot["A1"] = "Cirurgião"
    ws_tot["B1"] = "Total (R$)"
    ws_tot["C1"] = "Total Unimed"
    ws_tot["D1"] = "Diferença"  # ✅ C - B

    for c in ("A1", "B1", "C1", "D1"):
        ws_tot[c].font = Font(bold=True)

    # ranges para SUMIF (ignora a linha Total do Resumo)
    cir_range = f"Resumo!${cir_col_letter}$2:${cir_col_letter}${data_last_row}"
    tot_range = f"Resumo!${total_col_letter}$2:${total_col_letter}${data_last_row}"

    # mapa header -> letra no Resumo
    col_letter_by_name = {
        str(h): get_column_letter(i + 1)
        for i, h in enumerate(headers)
        if h is not None
    }

    for i, nome in enumerate(CIRURGIOES, start=2):
        ws_tot[f"A{i}"] = nome

        # B: Total por cirurgião (a partir do dropdown na aba Resumo)
        ws_tot[f"B{i}"] = f"=SUMIF({cir_range},A{i},{tot_range})"
        ws_tot[f"B{i}"].number_format = 'R$ #,##0.00'

        # C: Total Unimed (rodapé da coluna MCN/DBA/JBM/GRM/AGL no Resumo)
        if nome in COLUNAS_RESUMO:
            col_resumo = col_letter_by_name.get(nome)
            if col_resumo:
                ws_tot[f"C{i}"] = f"=Resumo!${col_resumo}${last_row_resumo}"
                ws_tot[f"C{i}"].number_format = 'R$ #,##0.00'
            else:
                ws_tot[f"C{i}"] = ""
        else:
            ws_tot[f"C{i}"] = ""

        # D: Diferença = C - B
        ws_tot[f"D{i}"] = f"=IFERROR(N(C{i})-N(B{i}),0)"
        ws_tot[f"D{i}"].number_format = 'R$ #,##0.00'



    # Linha total geral (só na coluna B faz sentido, mas você pode somar D também se quiser)
    last_tot_row = 1 + len(CIRURGIOES) + 1
    ws_tot[f"A{last_tot_row}"] = "Total Geral"
    ws_tot[f"A{last_tot_row}"].font = Font(bold=True)

    ws_tot[f"B{last_tot_row}"] = f"=SUM(B2:B{last_tot_row-1})"
    ws_tot[f"B{last_tot_row}"].number_format = 'R$ #,##0.00'
    ws_tot[f"B{last_tot_row}"].font = Font(bold=True)

    ws_tot[f"C{last_tot_row}"] = f"=SUM(C2:C{last_tot_row-1})"
    ws_tot[f"C{last_tot_row}"].number_format = 'R$ #,##0.00'
    ws_tot[f"C{last_tot_row}"].font = Font(bold=True)

    ws_tot[f"D{last_tot_row}"] = f"=SUM(D2:D{last_tot_row-1})"
    ws_tot[f"D{last_tot_row}"].number_format = 'R$ #,##0.00'
    ws_tot[f"D{last_tot_row}"].font = Font(bold=True)

    # -------------------------
    # Formatar moeda no Resumo (colunas numéricas, exceto Cirurgião)
    # -------------------------
    for col_idx in range(2, ws_resumo.max_column + 1):
        header = ws_resumo.cell(row=1, column=col_idx).value
        if header == "Cirurgião":
            continue
        col_letter = get_column_letter(col_idx)
        for cell in ws_resumo[col_letter][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'

    # Ajustar larguras
    autofit_worksheet(ws_resumo, min_width_numeric=18, currency_cols_from=2)
    autofit_worksheet(ws_tot, min_width_numeric=18, currency_cols_from=2)
    autofit_worksheet(ws_listas, min_width_numeric=10, currency_cols_from=99)

print(f"✅ Gerado: {SAIDA}")
